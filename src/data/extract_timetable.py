import openpyxl
import json
import re
import os

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "IFTT (W2025).xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "facultytimetable.json")

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

DAY_MAP = {
    'mon': 'Monday',
    'tue': 'Tuesday',
    'wed': 'Wednesday',
    'thu': 'Thursday',
    'fri': 'Friday',
    'sat': 'Saturday',
    'sun': 'Sunday',
    'monday': 'Monday',
    'tuesday': 'Tuesday',
    'wednesday': 'Wednesday',
    'thursday': 'Thursday',
    'friday': 'Friday',
}

def find_row_by_content(ws, search_text, max_row=20):
    """Find the row number containing the search text."""
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=5, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and search_text.lower() in cell.value.lower():
                return cell.row
    return None

def get_cell_value(ws, row, col):
    """Get cell value, handling merged cells."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    # Check if cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Return the value of the top-left cell of the merged range
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None

def normalize_time(time_str):
    """Clean up time string formatting."""
    if not time_str or not isinstance(time_str, str):
        return str(time_str) if time_str else ""
    # Strip whitespace
    time_str = time_str.strip()
    return time_str

def determine_type(cell_value):
    """Determine if a slot is Theory, Lab, or Break."""
    if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ''):
        return "Break"
    
    val = str(cell_value)
    
    # Check for Lab indicators (case insensitive)
    lab_patterns = [
        r'\(L\)',          # (L)
        r'\bLab\b',        # Lab as a word
        r'\bLAB\b',        # LAB as a word
        r'-L\b',           # -L at end of word
        r'-L\n',           # -L followed by newline
        r'-L\s',           # -L followed by space
        r'\bL\s*$',        # L at end
        r'- L\b',          # - L
    ]
    
    for pattern in lab_patterns:
        if re.search(pattern, val, re.IGNORECASE):
            return "Lab"
    
    return "Theory"

def parse_cell_content(cell_value):
    """Parse cell content to extract subject and class."""
    if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ''):
        return "Break", "Break"
    
    val = str(cell_value).strip()
    return val, val

def extract_faculty_data(ws, sheet_name):
    """Extract faculty data from a worksheet."""
    
    # Find faculty name row
    faculty_row = find_row_by_content(ws, 'Faculty Name')
    if faculty_row is None:
        print(f"WARNING: Could not find 'Faculty Name' in sheet '{sheet_name}', skipping...")
        return None
    
    # Extract faculty info
    faculty_name = ""
    contact = ""
    designation = ""
    
    for col in range(1, 15):
        val = get_cell_value(ws, faculty_row, col)
        if val and isinstance(val, str):
            val_lower = val.lower().strip()
            if 'faculty name' in val_lower:
                # The name is in the next non-empty cell
                for c2 in range(col+1, col+5):
                    v2 = get_cell_value(ws, faculty_row, c2)
                    if v2 and isinstance(v2, str) and v2.strip() and 'contact' not in v2.lower() and 'designation' not in v2.lower() and 'email' not in v2.lower():
                        faculty_name = v2.strip()
                        break
            elif 'contact' in val_lower:
                for c2 in range(col+1, col+3):
                    v2 = get_cell_value(ws, faculty_row, c2)
                    if v2 is not None:
                        if isinstance(v2, float):
                            contact = str(int(v2))
                        elif isinstance(v2, int):
                            contact = str(v2)
                        else:
                            contact = str(v2).strip()
                        break
            elif 'designation' in val_lower:
                for c2 in range(col+1, col+4):
                    v2 = get_cell_value(ws, faculty_row, c2)
                    if v2 and isinstance(v2, str) and v2.strip():
                        designation = v2.strip()
                        break
            elif 'email' in val_lower:
                # skip email fields
                pass
    
    # Find time header row (contains "Time" and "Day")
    time_row = find_row_by_content(ws, 'Time')
    if time_row is None:
        time_row = find_row_by_content(ws, 'Day')
    if time_row is None:
        print(f"WARNING: Could not find time header in sheet '{sheet_name}', skipping...")
        return None
    
    # Find period row
    period_row = find_row_by_content(ws, 'Period')
    if period_row is None:
        period_row = time_row + 1
    
    # Extract time slots from the time header row
    time_slots = {}
    max_col = ws.max_column
    if max_col > 20:
        max_col = 20
    
    for col in range(3, max_col + 1):
        val = get_cell_value(ws, time_row, col)
        if val and isinstance(val, str) and val.strip():
            time_slots[col] = normalize_time(val)
    
    # Extract period numbers
    periods = {}
    for col in range(3, max_col + 1):
        val = get_cell_value(ws, period_row, col)
        if val is not None:
            try:
                periods[col] = int(float(val))
            except (ValueError, TypeError):
                pass
    
    # Find day rows (Mon-Fri starting after period row)
    day_start_row = period_row + 1
    days_data = []
    
    for row_idx in range(day_start_row, day_start_row + 7):
        day_cell = get_cell_value(ws, row_idx, 1)
        if day_cell is None:
            continue
        day_str = str(day_cell).strip().lower()
        
        # Map to full day name
        full_day = None
        for key, value in DAY_MAP.items():
            if day_str.startswith(key) or key.startswith(day_str):
                full_day = value
                break
        
        if full_day is None:
            continue
        
        # Get allocated shift
        shift_val = get_cell_value(ws, row_idx, 2)
        allocated_shift = str(shift_val).strip() if shift_val else ""
        
        # Extract slots for this day
        slots = []
        for col in range(3, max_col + 1):
            if col not in time_slots and col not in periods:
                continue
            
            period_num = periods.get(col, col - 2)
            time_val = time_slots.get(col, "")
            
            cell_val = get_cell_value(ws, row_idx, col)
            
            slot_type = determine_type(cell_val)
            subject, class_name = parse_cell_content(cell_val)
            
            slot = {
                "period": period_num,
                "time": time_val,
                "subject": subject,
                "class": class_name,
                "type": slot_type
            }
            slots.append(slot)
        
        days_data.append({
            "day": full_day,
            "allocated_shift": allocated_shift,
            "slots": slots
        })
    
    return {
        "faculty_name": faculty_name,
        "faculty_designation": designation,
        "contact": contact,
        "timetable": days_data
    }

# Process all sheets
all_faculty = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Processing sheet: {sheet_name}")
    
    faculty_data = extract_faculty_data(ws, sheet_name)
    if faculty_data:
        all_faculty.append(faculty_data)
    else:
        print(f"  -> Skipped (no data found)")

# Write output
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(all_faculty, f, indent=2, ensure_ascii=False)

print(f"\nDone! Extracted {len(all_faculty)} faculty timetables to {OUTPUT_PATH}")
